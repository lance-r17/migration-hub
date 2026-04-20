import io
from collections import defaultdict
from decimal import Decimal

from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.models.billing_breakdown_record import BillingBreakdownRecord
from app.models.billing_record import BillingRecord
from app.models.config_store import ConfigStore
from app.schemas.billing import (
    BillingBreakdownRecordIn,
    BillingThresholdConfigOut,
    BillingThresholdConfigUpdate,
    BillingUpload,
)

_BILLING_THRESHOLD_KEY = "billing_threshold_config"
_DEFAULT_THRESHOLDS = {
    "healthyAtRiskThreshold": 100.0,
    "atRiskOverThreshold": 120.0,
    "currency": "CNY",
}

_SUMMARY_SHEET = "Resources Set Summary"
_BILLING_SHEET = "Resources Set Billing"


def parse_xlsx_report(
    file_bytes: bytes, month: str, env: str
) -> tuple[list[dict], list[BillingBreakdownRecordIn]]:
    """Parse an Alibaba Cloud billing xlsx report.

    Returns (summary_records, breakdown_records).
    Raises ValueError with a descriptive message if the file is invalid.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for xlsx parsing")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read xlsx file: {exc}") from exc

    if _SUMMARY_SHEET not in wb.sheetnames:
        raise ValueError(f'Missing sheet "{_SUMMARY_SHEET}"')
    if _BILLING_SHEET not in wb.sheetnames:
        raise ValueError(f'Missing sheet "{_BILLING_SHEET}"')

    # ── Parse "Resources Set Summary" ─────────────────────────────────────────
    ws_summary = wb[_SUMMARY_SHEET]
    summary_rows = list(ws_summary.iter_rows(values_only=True))
    if not summary_rows:
        raise ValueError(f'Sheet "{_SUMMARY_SHEET}" is empty')

    headers_s = [str(h).lower().strip() if h else '' for h in summary_rows[0]]
    # Col B (idx 1) = resource set; Col G (idx 6) = Final Amount
    if len(headers_s) < 7:
        raise ValueError(f'Sheet "{_SUMMARY_SHEET}" must have at least 7 columns')
    if 'resources set' not in headers_s[1] and 'resource' not in headers_s[1]:
        raise ValueError(
            f'Sheet "{_SUMMARY_SHEET}" column B must be the resource set name '
            f'(got "{summary_rows[0][1]}")'
        )

    summary_records: list[dict] = []
    for row in summary_rows[1:]:
        if not row or row[1] is None:
            continue
        resource_set = str(row[1]).strip()
        raw_amount = row[6] if len(row) > 6 else None
        if not resource_set or raw_amount is None:
            continue
        try:
            amount = float(raw_amount)
        except (TypeError, ValueError):
            continue
        summary_records.append({"resource_set": resource_set, "amount": amount})

    if not summary_records:
        raise ValueError(f'Sheet "{_SUMMARY_SHEET}" contains no valid data rows')

    # ── Parse "Resources Set Billing" ─────────────────────────────────────────
    ws_billing = wb[_BILLING_SHEET]
    billing_rows = list(ws_billing.iter_rows(values_only=True))
    if not billing_rows:
        raise ValueError(f'Sheet "{_BILLING_SHEET}" is empty')

    if len(billing_rows[0]) < 19:
        raise ValueError(
            f'Sheet "{_BILLING_SHEET}" must have at least 19 columns '
            f'(got {len(billing_rows[0])})'
        )

    # Col E(4)=resource_set, F(5)=product, S(18)=Final Amount
    aggregated: dict[tuple[str, str], float] = defaultdict(float)
    for row in billing_rows[1:]:
        if not row or row[4] is None:
            continue
        resource_set = str(row[4]).strip()
        product = str(row[5]).strip() if row[5] else ''
        raw_amount = row[18] if len(row) > 18 else None
        if not resource_set or not product or raw_amount is None:
            continue
        try:
            aggregated[(resource_set, product)] += float(raw_amount)
        except (TypeError, ValueError):
            continue

    breakdown_records = [
        BillingBreakdownRecordIn(resource_set=rs, product=prod, amount=round(amt, 2))
        for (rs, prod), amt in aggregated.items()
    ]

    wb.close()
    return summary_records, breakdown_records


async def get_months(session: AsyncSession, env: str) -> list[str]:
    result = await session.execute(
        select(BillingRecord.month).where(BillingRecord.env == env).distinct().order_by(BillingRecord.month)
    )
    return [r[0] for r in result.all()]


async def get_records(session: AsyncSession, month: str, env: str) -> list[BillingRecord]:
    result = await session.execute(
        select(BillingRecord).where(BillingRecord.month == month, BillingRecord.env == env)
    )
    return list(result.scalars().all())


async def upsert_records(session: AsyncSession, upload: BillingUpload) -> None:
    await session.execute(
        delete(BillingRecord).where(BillingRecord.month == upload.month, BillingRecord.env == upload.env)
    )
    for rec in upload.records:
        session.add(BillingRecord(
            month=upload.month,
            env=upload.env,
            resource_set=rec.resource_set,
            amount=rec.amount,
        ))
    await session.flush()


async def upsert_breakdown_records(
    session: AsyncSession, month: str, env: str, records: list[BillingBreakdownRecordIn]
) -> None:
    await session.execute(
        delete(BillingBreakdownRecord).where(
            BillingBreakdownRecord.month == month,
            BillingBreakdownRecord.env == env,
        )
    )
    for rec in records:
        session.add(BillingBreakdownRecord(
            month=month,
            env=env,
            resource_set=rec.resource_set,
            product=rec.product,
            amount=Decimal(str(rec.amount)),
        ))
    await session.flush()


async def get_breakdown_records(
    session: AsyncSession, month: str, env: str, resource_set: str
) -> list[BillingBreakdownRecord]:
    result = await session.execute(
        select(BillingBreakdownRecord).where(
            BillingBreakdownRecord.month == month,
            BillingBreakdownRecord.env == env,
            BillingBreakdownRecord.resource_set == resource_set,
        )
    )
    return list(result.scalars().all())


async def delete_month_records(session: AsyncSession, month: str) -> None:
    await session.execute(
        delete(BillingBreakdownRecord).where(BillingBreakdownRecord.month == month)
    )
    await session.execute(
        delete(BillingRecord).where(BillingRecord.month == month)
    )
    await session.flush()


async def get_threshold_config(session: AsyncSession) -> BillingThresholdConfigOut:
    row = await session.get(ConfigStore, _BILLING_THRESHOLD_KEY)
    data = row.value if row else _DEFAULT_THRESHOLDS
    return BillingThresholdConfigOut(
        healthy_at_risk_threshold=data.get("healthyAtRiskThreshold", 100.0),
        at_risk_over_threshold=data.get("atRiskOverThreshold", 120.0),
        currency=data.get("currency", "CNY"),
    )


async def update_threshold_config(
    session: AsyncSession, patch: BillingThresholdConfigUpdate
) -> BillingThresholdConfigOut:
    row = await session.get(ConfigStore, _BILLING_THRESHOLD_KEY)
    current = row.value if row else dict(_DEFAULT_THRESHOLDS)
    if patch.healthy_at_risk_threshold is not None:
        current["healthyAtRiskThreshold"] = patch.healthy_at_risk_threshold
    if patch.at_risk_over_threshold is not None:
        current["atRiskOverThreshold"] = patch.at_risk_over_threshold
    if patch.currency is not None:
        current["currency"] = patch.currency
    if row:
        row.value = current
        flag_modified(row, 'value')
    else:
        session.add(ConfigStore(key=_BILLING_THRESHOLD_KEY, value=current))
    await session.flush()
    return BillingThresholdConfigOut(
        healthy_at_risk_threshold=current["healthyAtRiskThreshold"],
        at_risk_over_threshold=current["atRiskOverThreshold"],
        currency=current.get("currency", "CNY"),
    )
